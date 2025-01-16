import streamlit as st
import pandas as pd
import openai
import json
import io
import base64
import numpy as np
import re
from docx import Document
from datetime import datetime
from openpyxl import Workbook

###############################################################################
# 1. Utility Functions
###############################################################################
def generate_sample_excel_template():
    """
    Creates an Excel file in memory with the columns:
    URL, H1, Meta Description
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Pages"

    ws["A1"] = "URL"
    ws["B1"] = "H1"
    ws["C1"] = "Meta Description"

    # Example row
    ws["A2"] = "https://example.com/sample"
    ws["B2"] = "Sample Title"
    ws["C2"] = "Sample Meta Description"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def get_download_link(file_bytes: bytes, filename: str, link_label: str) -> str:
    """
    Generate an HTML link for downloading file_bytes as `filename`.
    """
    b64 = base64.b64encode(file_bytes).decode()
    return f'<a href="data:application/octet-stream;base64,{b64}" download="{filename}">{link_label}</a>'

def compute_cosine_similarity(vec_a: np.ndarray, vec_b: np.ndarray)->float:
    """
    Cosine similarity between two vectors
    """
    if np.linalg.norm(vec_a)==0 or np.linalg.norm(vec_b)==0:
        return 0.0
    return float(np.dot(vec_a, vec_b)/(np.linalg.norm(vec_a)*np.linalg.norm(vec_b)))

###############################################################################
# 2. Word Doc Parsing
###############################################################################
def parse_word_doc_paragraphs_only(doc_bytes: bytes):
    """
    Parse `.docx`, skip headings, return a list of dicts:
    [ { "content": "...", "embedding": None, "sentences": None }, ...]
    """
    doc = Document(io.BytesIO(doc_bytes))
    paragraphs=[]
    for para in doc.paragraphs:
        text=para.text.strip()
        if not text:
            continue
        style_name=para.style.name.lower() if para.style else ""
        if "heading" not in style_name:
            paragraphs.append({"content": text})
    return paragraphs

###############################################################################
# 3. Embeddings for site pages & doc paragraphs
###############################################################################
def embed_text_batch(openai_api_key:str, texts:list[str], model="text-embedding-ada-002"):
    """
    Batch-embed multiple `texts` with OpenAI Embedding API
    """
    openai.api_key=openai_api_key
    try:
        resp=openai.Embedding.create(model=model, input=texts)
        emb_list=[item["embedding"] for item in resp["data"]]
        return emb_list
    except Exception as e:
        st.error(f"Error generating embeddings: {e}")
        return [[] for _ in texts]

def embed_site_pages(df:pd.DataFrame, openai_api_key:str, batch_size:int=10)->pd.DataFrame:
    """
    For each row, combine URL+H1+Meta Desc => embed => store in df["embedding"].
    """
    st.info("Embedding site pages in batches...")
    texts=[]
    for _,row in df.iterrows():
        combined=f"{row['URL']} {row['H1']} {row['Meta Description']}"
        texts.append(combined)

    n=len(texts)
    all_embeddings=[]
    total_batches=(n+batch_size-1)//batch_size
    bar=st.progress(0)
    label=st.empty()

    idx=0
    for batch_idx in range(total_batches):
        start=batch_idx*batch_size
        end=start+batch_size
        batch_texts=texts[start:end]
        batch_embeddings=embed_text_batch(openai_api_key, batch_texts)
        all_embeddings.extend(batch_embeddings)

        idx+=1
        bar.progress(int(idx/total_batches*100))
        label.write(f"Site pages batch {idx}/{total_batches}")

    df["embedding"]=all_embeddings
    return df

def embed_doc_paragraphs(paragraphs:list[dict], openai_api_key:str, batch_size:int=10)->list[dict]:
    """
    Embed doc paragraphs, plus their sentences if needed
    """
    st.info("Embedding doc paragraphs in batches...")
    texts=[p["content"] for p in paragraphs]
    n=len(texts)
    all_embeddings=[]
    total_batches=(n+batch_size-1)//batch_size
    bar=st.progress(0)
    label=st.empty()

    idx=0
    for b_idx in range(total_batches):
        start=b_idx*batch_size
        end=start+batch_size
        batch_texts=texts[start:end]
        batch_embeds=embed_text_batch(openai_api_key, batch_texts)
        all_embeddings.extend(batch_embeds)

        idx+=1
        bar.progress(int(idx/total_batches*100))
        label.write(f"Paragraph batch {idx}/{total_batches}")

    # store
    for i,emb in enumerate(all_embeddings):
        paragraphs[i]["embedding"]=emb

    # embed each paragraph's sentences if you want finer detail
    st.info("Embedding paragraph sentences too...")
    bar_sents=st.progress(0)
    for i,p in enumerate(paragraphs):
        bar_sents.progress(int((i+1)/n*100))
        raw_txt=p["content"].strip()
        sents=re.split(r'(?<=[.!?])\s+',raw_txt)
        sents=[s.strip() for s in sents if s.strip()]
        if not sents:
            p["sentences"]=[]
            continue
        s_embs=embed_text_batch(openai_api_key, sents)
        p["sentences"]=list(zip(sents,s_embs))

    return paragraphs

###############################################################################
# 4. GPT-based Internal Links (Brand-New or Existing)
###############################################################################
def build_prompt_for_mode(mode:str, topic:str, keyword:str, target_data:dict, candidate_pages:pd.DataFrame)->str:
    """
    Construct the GPT prompt. 
    """
    anchor_text_guidelines="""
Anchor Text Best Practices:
- Descriptive, concise, relevant
- Avoid 'click here', 'read more', or stuffing keywords
- Should read naturally
"""
    if mode=="brand_new":
        brand_new_msg="This is a brand-new topic/page that doesn't exist yet.\n"
        target_url="N/A"
        target_h1=f"(New) {topic}"
        target_meta=f"A future page about {topic}"
    else:
        brand_new_msg=""
        target_url=target_data["URL"]
        target_h1=target_data["H1"]
        target_meta=target_data["Meta Description"]

    pages_info=[]
    for _,row in candidate_pages.iterrows():
        pages_info.append({
            "URL": row["URL"],
            "H1": row["H1"],
            "Meta Description": row["Meta Description"]
        })

    prompt=f"""
You are an SEO assistant. The user wants internal link recommendations.

Mode: {mode}
Topic: {topic}
Keyword (optional): {keyword}

{brand_new_msg}

Target Page:
- URL: {target_url}
- H1: {target_h1}
- Meta Description: {target_meta}

Candidate Pages in JSON:
{json.dumps(pages_info, indent=2)}

{anchor_text_guidelines}

Return only valid JSON. No disclaimers or code fences.

Each object:
  "target_url": "URL for the FROM page"
  "anchor_text": "Proposed anchor text"
""".strip()
    return prompt

def generate_internal_links(openai_api_key:str, mode:str, topic:str, keyword:str,
                            target_data:dict, candidate_pages:pd.DataFrame)->list:
    """
    Calls GPT to get internal link recommendations, parse JSON, handle errors.
    """
    openai.api_key=openai_api_key
    prompt=build_prompt_for_mode(mode, topic, keyword, target_data, candidate_pages)

    try:
        resp=openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role":"system",
                    "content":"You are a helpful SEO assistant. Return valid JSON only."
                },
                {
                    "role":"user",
                    "content": prompt
                }
            ],
            max_tokens=1500,
            temperature=0.0
        )
        raw=resp.choices[0].message["content"].strip()
        if not raw:
            st.warning("GPT returned empty text. Possibly out of tokens or an error.")
            return []
        try:
            data=json.loads(raw)
            if isinstance(data, dict):
                data=[data]
            return data
        except json.JSONDecodeError:
            st.warning(f"GPT returned invalid JSON:\n{raw}")
            return []
    except Exception as e:
        st.error(f"Error calling GPT: {e}")
        return []

###############################################################################
# 5. Main Streamlit App
###############################################################################
def main():
    st.title("All-In-One Internal Link Recommender (3 Modes)")

    st.write("""
    **Modes**:
    1) **Brand-New Topic/Page** (GPT-based)  
    2) **Optimize Existing Page** (GPT-based)  
    3) **Analyze Word Document** (pure embeddings, paragraphs choose unique pages)  
    """)

    # Step 1: Sample template
    st.subheader("Step 1: (Optional) Download Sample Template")
    sample_xlsx=generate_sample_excel_template()
    dl_link=get_download_link(sample_xlsx,"sample_template.xlsx","Download Sample XLSX")
    st.markdown(dl_link, unsafe_allow_html=True)

    # Step 2: Upload site pages
    st.subheader("Step 2: Upload Site Pages CSV/Excel")
    up_file=st.file_uploader("Columns: URL, H1, Meta Description", type=["csv","xlsx"])
    df=None
    if up_file:
        try:
            if up_file.name.endswith(".csv"):
                df=pd.read_csv(up_file)
            else:
                df=pd.read_excel(up_file)
        except Exception as e:
            st.error(f"Error reading pages: {e}")

    # Step 3: Mode
    st.subheader("Step 3: Choose Mode")
    mode_opts=["Brand-New Topic/Page","Optimize Existing Page","Analyze Word Document"]
    mode_choice=st.radio("Mode:", mode_opts)

    selected_url=None
    doc_paragraphs=[]
    if mode_choice=="Optimize Existing Page" and df is not None:
        needed_cols={"URL","H1","Meta Description"}
        if needed_cols.issubset(df.columns):
            all_urls=df["URL"].unique().tolist()
            selected_url=st.selectbox("Select Existing Page URL", all_urls)
        else:
            st.warning("Missing columns in your CSV/Excel.")
    elif mode_choice=="Analyze Word Document":
        st.subheader("3A: Upload Word Document (.docx)")
        doc_file=st.file_uploader("Upload .docx", type=["docx"])
        if doc_file:
            try:
                doc_bytes=doc_file.read()
                doc_paragraphs=parse_word_doc_paragraphs_only(doc_bytes)
                st.success(f"Found {len(doc_paragraphs)} paragraphs.")
            except Exception as e:
                st.error(f"Error parsing doc: {e}")

    # Step 4: For brand-new/existing, we get topic+keyword
    if mode_choice in ["Brand-New Topic/Page","Optimize Existing Page"]:
        st.subheader("Step 4: Enter Topic & Optional Keyword")
        topic=st.text_input("Topic (required)")
        keyword=st.text_input("Keyword (optional)")
    else:
        topic=""
        keyword=""

    # Step 5: OpenAI Key
    st.subheader("Step 5: Enter OpenAI API Key")
    openai_key=st.text_input("OpenAI Key", type="password")

    # Step 6: Batch size
    st.subheader("Step 6: Set Batch Size for Embedding")
    batch_sz=st.slider("Batch size",1,50,10)

    # Step 7: Embed
    st.subheader("Step 7: Embed Data")
    if st.button("Embed Now"):
        if not openai_key:
            st.error("Please provide an OpenAI key.")
            st.stop()
        if df is None or df.empty:
            st.error("Please upload site pages first or ensure file isn't empty.")
            st.stop()
        needed_cols={"URL","H1","Meta Description"}
        if not needed_cols.issubset(df.columns):
            st.error(f"File missing columns: {', '.join(needed_cols)}")
            st.stop()

        df_emb=embed_site_pages(df.copy(), openai_key,batch_sz)
        doc_data=None
        if mode_choice=="Analyze Word Document":
            if not doc_paragraphs:
                st.error("No paragraphs found in doc.")
                st.stop()
            doc_data=embed_doc_paragraphs(doc_paragraphs, openai_key,batch_sz)

        # Store everything
        st.session_state["df_pages"]=df_emb
        st.session_state["doc_data"]=doc_data
        st.session_state["mode"]=mode_choice
        st.session_state["topic"]=topic
        st.session_state["keyword"]=keyword
        st.session_state["selected_url"]=selected_url
        st.success("Embedding complete. Proceed to Generate Links below.")

    # Step 8: Generate
    st.subheader("Step 8: Generate Links")
    if st.button("Generate Links"):
        if "df_pages" not in st.session_state:
            st.error("Please embed data first.")
            st.stop()

        final_mode=st.session_state["mode"]
        df_pages_emb=st.session_state["df_pages"]
        doc_data=st.session_state["doc_data"]
        final_topic=st.session_state["topic"]
        final_keyword=st.session_state["keyword"]
        final_url=st.session_state.get("selected_url", None)

        threshold=0.80

        if final_mode in ["Brand-New Topic/Page","Optimize Existing Page"]:
            # GPT-based approach
            if final_mode=="Optimize Existing Page":
                try:
                    target_row=df_pages_emb.loc[df_pages_emb["URL"]==final_url].iloc[0].to_dict()
                except IndexError:
                    st.error("Selected URL not found in dataset.")
                    st.stop()
                candidate_df=df_pages_emb.loc[df_pages_emb["URL"]!=final_url].copy()

                # If user typed a topic => do a quick similarity filter
                if final_topic:
                    from_embed=embed_text_batch(openai_key,[final_topic+" "+final_keyword if final_keyword else final_topic])
                    if from_embed and from_embed[0]:
                        q_vec=np.array(from_embed[0])
                        sims=[]
                        for i,row in candidate_df.iterrows():
                            page_vec=np.array(row["embedding"])
                            sim_val=compute_cosine_similarity(q_vec,page_vec)
                            sims.append(sim_val)
                        candidate_df["similarity"]=sims
                        candidate_df=candidate_df[candidate_df["similarity"]>=threshold]

                links=generate_internal_links(
                    openai_key,
                    "existing_page",
                    final_topic,
                    final_keyword,
                    target_row,
                    candidate_df[["URL","H1","Meta Description","similarity"]] if "similarity" in candidate_df.columns else candidate_df[["URL","H1","Meta Description"]]
                )
                if not links:
                    st.warning("No recommendations or GPT returned invalid JSON.")
                else:
                    st.subheader("Optimize Existing Page Results")
                    for item in links:
                        anchor_txt=item.get("anchor_text","(No anchor)")
                        page_url=item.get("target_url","")
                        st.markdown(f"- **Page Link**: [{anchor_txt}]({page_url})")
                        # If we stored 'similarity'
                        sim_val=item.get("similarity",None)
                        if sim_val is not None:
                            st.markdown(f"  - **Similarity Score**: {round(sim_val*100,2)}%")

            else:
                # brand_new
                if final_topic:
                    q_embed=embed_text_batch(openai_key,[final_topic+" "+final_keyword if final_keyword else final_topic])
                    if q_embed and q_embed[0]:
                        q_vec=np.array(q_embed[0])
                        sims=[]
                        for i,row in df_pages_emb.iterrows():
                            page_vec=np.array(row["embedding"])
                            sim_val=compute_cosine_similarity(q_vec,page_vec)
                            sims.append(sim_val)
                        df_pages_emb["similarity"]=sims
                        df_pages_emb=df_pages_emb[df_pages_emb["similarity"]>=threshold]

                new_target={
                    "URL":"N/A",
                    "H1":f"(New) {final_topic}",
                    "Meta Description":f"Future page about {final_topic}"
                }
                links=generate_internal_links(
                    openai_key,
                    "brand_new",
                    final_topic,
                    final_keyword,
                    new_target,
                    df_pages_emb[["URL","H1","Meta Description","similarity"]] if "similarity" in df_pages_emb.columns else df_pages_emb[["URL","H1","Meta Description"]]
                )
                if not links:
                    st.warning("No recommendations or GPT returned invalid JSON.")
                else:
                    st.subheader("Brand-New Topic Results")
                    for item in links:
                        anchor_txt=item.get("anchor_text","(No anchor)")
                        page_url=item.get("target_url","")
                        st.markdown(f"- **Page Link**: [{anchor_txt}]({page_url})")
                        sim_val=item.get("similarity",None)
                        if sim_val is not None:
                            st.markdown(f"  - **Similarity Score**: {round(sim_val*100,2)}%")

        else:
            # Word Doc => each paragraph picks exactly 1 page (≥80%), removing page from pool
            if not doc_data:
                st.error("No doc data found. Please embed the doc paragraphs.")
                st.stop()

            doc_df=pd.DataFrame(doc_data)
            pages_copy=df_pages_emb.copy()
            results=[]
            paragraphs_count=len(doc_df)

            st.write("**Inverse Approach**: each paragraph picks 1 unique page (≥80% sim), that page is removed from the pool.")
            bar=st.progress(0)
            label=st.empty()

            for p_idx, p_row in doc_df.iterrows():
                progress_val=int(((p_idx+1)/paragraphs_count)*100)
                bar.progress(progress_val)
                label.write(f"Paragraph {p_idx+1}/{paragraphs_count}")

                para_vec=np.array(p_row["embedding"])
                best_sim=-1.0
                best_idx=-1
                for i, page_row in pages_copy.iterrows():
                    page_vec=np.array(page_row["embedding"])
                    sim_val=compute_cosine_similarity(para_vec,page_vec)
                    if sim_val>best_sim:
                        best_sim=sim_val
                        best_idx=i
                if best_sim>=0.80 and best_idx!=-1:
                    chosen_page=pages_copy.loc[best_idx]
                    results.append({
                        "paragraph_index": p_idx,
                        "paragraph_text": p_row["content"],
                        "page_url": chosen_page["URL"],
                        "page_title": chosen_page["H1"],
                        "similarity_score": best_sim
                    })
                    pages_copy.drop(index=best_idx,inplace=True)
                if pages_copy.empty:
                    break

            if not results:
                st.warning("No paragraphs matched any page above 80%, or no pages left.")
            else:
                st.subheader("Word Doc Results")
                for item in results:
                    with st.expander(f"Paragraph #{item['paragraph_index']+1}"):
                        st.markdown(f"> **Paragraph Text**:\n> {item['paragraph_text']}")
                        st.markdown(f"- **Page Link**: [{item['page_title']}]({item['page_url']})")
                        st.markdown(f"- **Similarity Score**: {round(item['similarity_score']*100,2)}%")

                df_final=pd.DataFrame(results)
                st.subheader("Download CSV")
                csv_data=df_final.to_csv(index=False).encode("utf-8")
                st.download_button("Download CSV", csv_data, file_name=f"doc_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",mime="text/csv")

if __name__=="__main__":
    main()
